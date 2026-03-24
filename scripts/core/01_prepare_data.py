from __future__ import annotations

"""
Stage 1: Prepare and clean source datasets into canonical tabular files.

Inputs (raw):
- data/raw/tmd/dataset_halfSecondWindow.json
- data/raw/bridge/dataset_*.xlsx OR dataset_*.csv + rawdata_acceleration_*.csv
- data/raw/uci_har/*

Outputs (interim):
- data/interim/tmd_clean_mapped.csv
- data/interim/bridge_mapped.csv
- data/interim/uci_mapped.csv
- data/interim/all_sources_canonical.csv
"""

import argparse
import bisect
import csv
import json
import math
import struct
from datetime import datetime
from pathlib import Path


TMD_LABEL_MAP = {
    "walking": "human",
    "car": "vehicle",
    "bus": "vehicle",
    "still": "noise",
}

UCI_LABEL_MAP = {
    1: "human",  # WALKING
    2: "human",  # WALKING_UPSTAIRS
    3: "human",  # WALKING_DOWNSTAIRS
    4: "noise",  # SITTING
    5: "noise",  # STANDING
    6: "noise",  # LAYING
}

# Keep only motion features that are useful for this project.
CORE_FEATURES = [
    "android.sensor.accelerometer#mean",
    "android.sensor.accelerometer#min",
    "android.sensor.accelerometer#max",
    "android.sensor.accelerometer#std",
    "android.sensor.linear_acceleration#mean",
    "android.sensor.linear_acceleration#min",
    "android.sensor.linear_acceleration#max",
    "android.sensor.linear_acceleration#std",
    "android.sensor.gyroscope#mean",
    "android.sensor.gyroscope#min",
    "android.sensor.gyroscope#max",
    "android.sensor.gyroscope#std",
    "android.sensor.gyroscope_uncalibrated#mean",
    "android.sensor.gyroscope_uncalibrated#min",
    "android.sensor.gyroscope_uncalibrated#max",
    "android.sensor.gyroscope_uncalibrated#std",
]

HEADER = [
    "source",
    "sample_id",
    "user",
    "raw_target",
    "target",
    "time_sec",
    *CORE_FEATURES,
]


def to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return v if math.isfinite(v) else None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            v = float(s)
        except ValueError:
            return None
        return v if math.isfinite(v) else None
    return None


def normalize_label(raw: str) -> str:
    return raw.strip().lower().replace("-", "_").replace(" ", "_")


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def stats(values: list[float]) -> tuple[float, float, float, float]:
    if not values:
        return 0.0, 0.0, 0.0, 0.0
    n = len(values)
    mean_v = sum(values) / n
    min_v = min(values)
    max_v = max(values)
    var = sum((v - mean_v) * (v - mean_v) for v in values) / n
    std_v = math.sqrt(var) if var > 0 else 0.0
    return mean_v, min_v, max_v, std_v


def write_header(path: Path) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADER)
        writer.writeheader()


def append_rows(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    ensure_parent(path)
    with path.open("a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADER)
        writer.writerows(rows)


def parse_window_line(line: str) -> list[float]:
    parts = line.strip().split()
    out: list[float] = []
    for p in parts:
        try:
            out.append(float(p))
        except ValueError:
            continue
    return out


def parse_bridge_datetime(text: str) -> datetime | None:
    s = text.strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def load_bridge_label_map(path: Path) -> dict[datetime, str]:
    mapping: dict[datetime, str] = {}
    if not path.exists():
        return mapping

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            dt_text = (row.get("datetime") or "").strip()
            tgt = (row.get("target") or "").strip().lower()
            dt = parse_bridge_datetime(dt_text)
            if dt is None or tgt not in {"human", "vehicle", "noise"}:
                continue
            mapping[dt] = tgt
    return mapping


def write_bridge_label_template(bridge_dir: Path, out_path: Path) -> int:
    seen: set[str] = set()
    for p in sorted(bridge_dir.glob("dataset_*.csv")):
        with p.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            _ = next(reader, None)
            for row in reader:
                if not row:
                    continue
                dt = row[0].strip()
                if dt:
                    seen.add(dt)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["datetime", "target"])
        for dt in sorted(seen):
            writer.writerow([dt, ""])
    return len(seen)


def load_raw_accel_csv(path: Path) -> tuple[datetime | None, list[float], list[float], list[float], list[float]]:
    start_dt: datetime | None = None
    t_vals: list[float] = []
    x_vals: list[float] = []
    y_vals: list[float] = []
    z_vals: list[float] = []

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        first = next(reader, None)
        if first and first:
            start_dt = parse_bridge_datetime(str(first[0]))

        for row in reader:
            if len(row) < 4:
                continue
            t = to_float(row[0])
            x = to_float(row[1])
            y = to_float(row[2])
            z = to_float(row[3])
            if t is None or x is None or y is None or z is None:
                continue
            t_vals.append(t)
            x_vals.append(x)
            y_vals.append(y)
            z_vals.append(z)

    return start_dt, t_vals, x_vals, y_vals, z_vals


def magnitude3(ax: list[float], ay: list[float], az: list[float]) -> list[float]:
    n = min(len(ax), len(ay), len(az))
    return [math.sqrt(ax[i] * ax[i] + ay[i] * ay[i] + az[i] * az[i]) for i in range(n)]


def extract_sampled_float_stats(path: Path, max_floats: int = 8192) -> tuple[float, float, float, float] | None:
    """
    Fast, dependency-free approximate waveform stats.
    Reads a small float32 sample from the file and computes |signal| stats.
    """
    try:
        raw = path.read_bytes()
    except Exception:
        return None

    if len(raw) < 4096:
        return None

    # SAC commonly has a 632-byte header; for .sac-hd we also try from 0 fallback.
    offsets = [632, 0] if path.suffix.lower() == ".sac" else [0, 632]
    vals: list[float] = []

    for off in offsets:
        if off >= len(raw):
            continue
        available = len(raw) - off
        n = min(max_floats, available // 4)
        if n <= 0:
            continue

        try:
            chunk = raw[off : off + (n * 4)]
            unpacked = struct.unpack("<" + ("f" * n), chunk)
        except Exception:
            continue

        vals = [abs(v) for v in unpacked if math.isfinite(v) and abs(v) < 1e7]
        if len(vals) >= 64:
            break

    if len(vals) < 64:
        return None

    return stats(vals)


def prepare_tmd(data_root: Path, progress_every: int) -> tuple[int, int, dict[str, int]]:
    tmd_path = data_root / "raw" / "tmd" / "dataset_halfSecondWindow.json"
    out_tmd = data_root / "interim" / "tmd_clean_mapped.csv"
    out_all = data_root / "interim" / "all_sources_canonical.csv"

    if not tmd_path.exists():
        raise FileNotFoundError(f"Missing TMD input file: {tmd_path}")

    ensure_parent(out_tmd)
    ensure_parent(out_all)

    kept = 0
    dropped = 0
    class_counts: dict[str, int] = {"human": 0, "vehicle": 0, "noise": 0}

    print(f"[TMD] Reading: {tmd_path}", flush=True)
    with tmd_path.open("r", encoding="utf-8") as fin, out_tmd.open("w", encoding="utf-8", newline="") as ftmd, out_all.open("a", encoding="utf-8", newline="") as fall:
        tmd_writer = csv.DictWriter(ftmd, fieldnames=HEADER)
        all_writer = csv.DictWriter(fall, fieldnames=HEADER)
        tmd_writer.writeheader()

        seen = 0
        for line in fin:
            seen += 1
            line = line.strip()
            if not line:
                continue

            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                dropped += 1
                continue

            if not isinstance(row, dict):
                dropped += 1
                continue

            raw_target = str(row.get("target", "")).strip()
            mapped = TMD_LABEL_MAP.get(normalize_label(raw_target))
            if not mapped:
                dropped += 1
                continue

            values: dict[str, float] = {}
            invalid = False
            for feat in CORE_FEATURES:
                v = to_float(row.get(feat))
                if v is None:
                    invalid = True
                    break
                # Basic sanity clipping guard: remove clearly corrupted numeric rows.
                if abs(v) > 1_000_000:
                    invalid = True
                    break
                values[feat] = v

            if invalid:
                dropped += 1
                continue

            time_sec = to_float(row.get("time"))
            out_row: dict[str, str] = {
                "source": "tmd",
                "sample_id": str(row.get("id", "")),
                "user": str(row.get("user", "")),
                "raw_target": raw_target,
                "target": mapped,
                "time_sec": "" if time_sec is None else f"{time_sec:.6f}",
            }
            for feat in CORE_FEATURES:
                out_row[feat] = f"{values[feat]:.8f}"

            tmd_writer.writerow(out_row)
            all_writer.writerow(out_row)
            kept += 1
            class_counts[mapped] += 1

            if progress_every > 0 and seen % progress_every == 0:
                print(f"[TMD] processed={seen} kept={kept} dropped={dropped}", flush=True)

    return kept, dropped, class_counts


def prepare_uci_har(data_root: Path, progress_every: int) -> tuple[int, int, dict[str, int]]:
    uci_dir = data_root / "raw" / "uci_har"
    out_uci = data_root / "interim" / "uci_mapped.csv"
    out_all = data_root / "interim" / "all_sources_canonical.csv"

    if not uci_dir.exists():
        write_placeholder_csv(out_uci)
        return 0, 0, {"human": 0, "vehicle": 0, "noise": 0}

    rows: list[dict[str, str]] = []
    class_counts = {"human": 0, "vehicle": 0, "noise": 0}
    dropped = 0

    print(f"[UCI] Reading: {uci_dir}", flush=True)
    for split in ["train", "test"]:
        inertial = uci_dir / split / "Inertial Signals"
        y_path = uci_dir / split / f"y_{split}.txt"
        subj_path = uci_dir / split / f"subject_{split}.txt"

        tx = inertial / f"total_acc_x_{split}.txt"
        ty = inertial / f"total_acc_y_{split}.txt"
        tz = inertial / f"total_acc_z_{split}.txt"
        bx = inertial / f"body_acc_x_{split}.txt"
        by = inertial / f"body_acc_y_{split}.txt"
        bz = inertial / f"body_acc_z_{split}.txt"
        gx = inertial / f"body_gyro_x_{split}.txt"
        gy = inertial / f"body_gyro_y_{split}.txt"
        gz = inertial / f"body_gyro_z_{split}.txt"

        needed = [y_path, subj_path, tx, ty, tz, bx, by, bz, gx, gy, gz]
        if not all(p.exists() for p in needed):
            continue

        with y_path.open("r", encoding="utf-8") as fy, subj_path.open("r", encoding="utf-8") as fs, tx.open("r", encoding="utf-8") as ftx, ty.open("r", encoding="utf-8") as fty, tz.open("r", encoding="utf-8") as ftz, bx.open("r", encoding="utf-8") as fbx, by.open("r", encoding="utf-8") as fby, bz.open("r", encoding="utf-8") as fbz, gx.open("r", encoding="utf-8") as fgx, gy.open("r", encoding="utf-8") as fgy, gz.open("r", encoding="utf-8") as fgz:
            idx = 0
            print(f"[UCI] processing split={split}", flush=True)
            for y_line in fy:
                idx += 1
                raw_label = y_line.strip()
                try:
                    y_id = int(raw_label)
                except ValueError:
                    dropped += 1
                    continue

                mapped = UCI_LABEL_MAP.get(y_id)
                if not mapped:
                    dropped += 1
                    continue

                user = fs.readline().strip()
                total_mag = magnitude3(
                    parse_window_line(ftx.readline()),
                    parse_window_line(fty.readline()),
                    parse_window_line(ftz.readline()),
                )
                body_mag = magnitude3(
                    parse_window_line(fbx.readline()),
                    parse_window_line(fby.readline()),
                    parse_window_line(fbz.readline()),
                )
                gyro_mag = magnitude3(
                    parse_window_line(fgx.readline()),
                    parse_window_line(fgy.readline()),
                    parse_window_line(fgz.readline()),
                )

                if not total_mag or not body_mag or not gyro_mag:
                    dropped += 1
                    continue

                a_mean, a_min, a_max, a_std = stats(total_mag)
                l_mean, l_min, l_max, l_std = stats(body_mag)
                g_mean, g_min, g_max, g_std = stats(gyro_mag)

                row = {
                    "source": "uci_har",
                    "sample_id": f"{split}_{idx}",
                    "user": user,
                    "raw_target": str(y_id),
                    "target": mapped,
                    "time_sec": "2.56",
                    "android.sensor.accelerometer#mean": f"{a_mean:.8f}",
                    "android.sensor.accelerometer#min": f"{a_min:.8f}",
                    "android.sensor.accelerometer#max": f"{a_max:.8f}",
                    "android.sensor.accelerometer#std": f"{a_std:.8f}",
                    "android.sensor.linear_acceleration#mean": f"{l_mean:.8f}",
                    "android.sensor.linear_acceleration#min": f"{l_min:.8f}",
                    "android.sensor.linear_acceleration#max": f"{l_max:.8f}",
                    "android.sensor.linear_acceleration#std": f"{l_std:.8f}",
                    "android.sensor.gyroscope#mean": f"{g_mean:.8f}",
                    "android.sensor.gyroscope#min": f"{g_min:.8f}",
                    "android.sensor.gyroscope#max": f"{g_max:.8f}",
                    "android.sensor.gyroscope#std": f"{g_std:.8f}",
                    # UCI has calibrated gyro only; mirror into uncalibrated slots for schema consistency.
                    "android.sensor.gyroscope_uncalibrated#mean": f"{g_mean:.8f}",
                    "android.sensor.gyroscope_uncalibrated#min": f"{g_min:.8f}",
                    "android.sensor.gyroscope_uncalibrated#max": f"{g_max:.8f}",
                    "android.sensor.gyroscope_uncalibrated#std": f"{g_std:.8f}",
                }
                rows.append(row)
                class_counts[mapped] += 1

                if progress_every > 0 and idx % progress_every == 0:
                    print(f"[UCI] split={split} processed={idx} kept={len(rows)} dropped={dropped}", flush=True)

    ensure_parent(out_uci)
    with out_uci.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADER)
        writer.writeheader()
        writer.writerows(rows)

    append_rows(out_all, rows)
    return len(rows), dropped, class_counts


def prepare_bridge(
    data_root: Path,
    progress_every: int,
    bridge_label_map_path: Path | None = None,
) -> tuple[int, int, dict[str, int], str | None]:
    bridge_dir = data_root / "raw" / "bridge"
    out_bridge = data_root / "interim" / "bridge_mapped.csv"
    out_all = data_root / "interim" / "all_sources_canonical.csv"

    if not bridge_dir.exists():
        write_placeholder_csv(out_bridge)
        return 0, 0, {"human": 0, "vehicle": 0, "noise": 0}, "bridge input folder not found"

    rows: list[dict[str, str]] = []
    class_counts = {"human": 0, "vehicle": 0, "noise": 0}
    dropped = 0
    note: str | None = None

    def _print_progress(prefix: str, current: int, total: int) -> None:
        if total <= 0:
            return
        width = 28
        filled = int(width * min(max(current / total, 0.0), 1.0))
        bar = "#" * filled + "-" * (width - filled)
        pct = 100.0 * current / total
        print(f"\r{prefix} [{bar}] {current}/{total} ({pct:5.1f}%)", end="", flush=True)

    def _push_bridge_row(sample_id: str, raw_target: str, target: str, time_sec: str, acc_mag: list[float]) -> None:
        a_mean, a_min, a_max, a_std = stats(acc_mag)
        l_mean, l_min, l_max, l_std = a_mean, a_min, a_max, a_std
        g_mean, g_min, g_max, g_std = 0.0, 0.0, 0.0, 0.0

        row = {
            "source": "bridge",
            "sample_id": sample_id,
            "user": "bridge_site",
            "raw_target": raw_target,
            "target": target,
            "time_sec": time_sec,
            "android.sensor.accelerometer#mean": f"{a_mean:.8f}",
            "android.sensor.accelerometer#min": f"{a_min:.8f}",
            "android.sensor.accelerometer#max": f"{a_max:.8f}",
            "android.sensor.accelerometer#std": f"{a_std:.8f}",
            "android.sensor.linear_acceleration#mean": f"{l_mean:.8f}",
            "android.sensor.linear_acceleration#min": f"{l_min:.8f}",
            "android.sensor.linear_acceleration#max": f"{l_max:.8f}",
            "android.sensor.linear_acceleration#std": f"{l_std:.8f}",
            "android.sensor.gyroscope#mean": f"{g_mean:.8f}",
            "android.sensor.gyroscope#min": f"{g_min:.8f}",
            "android.sensor.gyroscope#max": f"{g_max:.8f}",
            "android.sensor.gyroscope#std": f"{g_std:.8f}",
            "android.sensor.gyroscope_uncalibrated#mean": f"{g_mean:.8f}",
            "android.sensor.gyroscope_uncalibrated#min": f"{g_min:.8f}",
            "android.sensor.gyroscope_uncalibrated#max": f"{g_max:.8f}",
            "android.sensor.gyroscope_uncalibrated#std": f"{g_std:.8f}",
        }
        rows.append(row)
        class_counts[target] += 1

    xlsx_files = sorted(bridge_dir.glob("dataset_*.xlsx"))
    if xlsx_files:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            write_placeholder_csv(out_bridge)
            return 0, 0, {"human": 0, "vehicle": 0, "noise": 0}, "openpyxl not installed; bridge ingest skipped"

        print(f"[Bridge] Found {len(xlsx_files)} workbook(s) in {bridge_dir}", flush=True)
        for xlsx in xlsx_files:
            print(f"[Bridge] loading {xlsx.name}", flush=True)
            try:
                wb = load_workbook(filename=str(xlsx), data_only=True, read_only=True)
            except Exception as e:
                dropped += 1
                print(f"[Bridge] skipped {xlsx.name} (load error: {e})", flush=True)
                continue
            if "label" not in wb.sheetnames or "sensorA_x" not in wb.sheetnames or "sensorA_y" not in wb.sheetnames or "sensorA_z" not in wb.sheetnames:
                wb.close()
                continue

            ws_label = wb["label"]
            ws_x = wb["sensorA_x"]
            ws_y = wb["sensorA_y"]
            ws_z = wb["sensorA_z"]
            max_row = min(ws_label.max_row, ws_x.max_row, ws_y.max_row, ws_z.max_row)

            label_rows = ws_label.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=4, values_only=True)
            x_rows = ws_x.iter_rows(min_row=2, max_row=max_row, values_only=True)
            y_rows = ws_y.iter_rows(min_row=2, max_row=max_row, values_only=True)
            z_rows = ws_z.iter_rows(min_row=2, max_row=max_row, values_only=True)
            total_rows = max(0, max_row - 1)

            for idx, (label_row, x_row, y_row, z_row) in enumerate(zip(label_rows, x_rows, y_rows, z_rows), start=1):
                human_cnt = to_float(label_row[0] if len(label_row) > 0 else None) or 0.0
                traina_cnt = to_float(label_row[1] if len(label_row) > 1 else None) or 0.0
                trainb_cnt = to_float(label_row[2] if len(label_row) > 2 else None) or 0.0
                car_cnt = to_float(label_row[3] if len(label_row) > 3 else None) or 0.0

                vehicle_cnt = traina_cnt + trainb_cnt + car_cnt
                if vehicle_cnt > 0 and human_cnt > 0:
                    dropped += 1
                    continue
                if vehicle_cnt > 0:
                    target = "vehicle"
                    raw_target = "vehicle_event"
                elif human_cnt > 0:
                    target = "human"
                    raw_target = "human_event"
                else:
                    target = "noise"
                    raw_target = "background"

                x_vals: list[float] = []
                y_vals: list[float] = []
                z_vals: list[float] = []
                max_col = min(len(x_row), len(y_row), len(z_row))
                for c in range(max_col):
                    xv = to_float(x_row[c])
                    yv = to_float(y_row[c])
                    zv = to_float(z_row[c])
                    if xv is None or yv is None or zv is None:
                        continue
                    x_vals.append(xv)
                    y_vals.append(yv)
                    z_vals.append(zv)

                if not x_vals or not y_vals or not z_vals:
                    dropped += 1
                    continue

                acc_mag = magnitude3(x_vals, y_vals, z_vals)
                if not acc_mag:
                    dropped += 1
                    continue

                _push_bridge_row(
                    sample_id=f"{xlsx.stem}_{idx + 1}",
                    raw_target=raw_target,
                    target=target,
                    time_sec="",
                    acc_mag=acc_mag,
                )

                if progress_every > 0 and idx % progress_every == 0:
                    _print_progress(f"[Bridge:{xlsx.name}]", idx, total_rows)
            if total_rows > 0:
                _print_progress(f"[Bridge:{xlsx.name}]", total_rows, total_rows)
                print("", flush=True)
            print(f"[Bridge] file={xlsx.name} kept_total={len(rows)} dropped_total={dropped}", flush=True)
            wb.close()
    else:
        event_files = sorted(bridge_dir.glob("dataset_*.csv"))
        raw_map = {
            p.name.replace("rawdata_acceleration_", "dataset_"): p
            for p in sorted(bridge_dir.glob("rawdata_acceleration_*.csv"))
        }
        if not event_files:
            write_placeholder_csv(out_bridge)
            return 0, 0, {"human": 0, "vehicle": 0, "noise": 0}, "no bridge dataset files found"

        if bridge_label_map_path is None:
            bridge_label_map_path = data_root.parent / "configs" / "bridge_event_labels.csv"
        label_map = load_bridge_label_map(bridge_label_map_path)

        if bridge_label_map_path and not label_map:
            n = write_bridge_label_template(bridge_dir, bridge_label_map_path)
            ensure_parent(out_bridge)
            with out_bridge.open("w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=HEADER)
                writer.writeheader()
            return 0, 0, class_counts, (
                f"bridge CSV detected: generated template with {n} events at {bridge_label_map_path}; "
                "fill target column with human/vehicle/noise"
            )

        print(f"[Bridge] CSV mode with {len(event_files)} event file(s)", flush=True)
        unlabeled = 0
        for event_csv in event_files:
            raw_csv = raw_map.get(event_csv.name)
            if raw_csv is None:
                dropped += 1
                continue

            start_dt, t_vals, x_vals, y_vals, z_vals = load_raw_accel_csv(raw_csv)
            if start_dt is None or not t_vals:
                dropped += 1
                continue

            with event_csv.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                _ = next(reader, None)
                for i, rec in enumerate(reader, start=1):
                    if not rec:
                        continue
                    dt_text = rec[0].strip()
                    ev_dt = parse_bridge_datetime(dt_text)
                    if ev_dt is None:
                        dropped += 1
                        continue

                    mapped = label_map.get(ev_dt)
                    if mapped not in {"human", "vehicle", "noise"}:
                        unlabeled += 1
                        dropped += 1
                        continue

                    offset = (ev_dt - start_dt).total_seconds()
                    l = bisect.bisect_left(t_vals, offset - 0.5)
                    r = bisect.bisect_right(t_vals, offset + 0.5)
                    if r - l < 3:
                        dropped += 1
                        continue

                    acc_mag = magnitude3(x_vals[l:r], y_vals[l:r], z_vals[l:r])
                    if not acc_mag:
                        dropped += 1
                        continue

                    _push_bridge_row(
                        sample_id=f"{event_csv.stem}_{i}",
                        raw_target=dt_text,
                        target=mapped,
                        time_sec=f"{offset:.3f}",
                        acc_mag=acc_mag,
                    )

                    if progress_every > 0 and i % progress_every == 0:
                        print(f"[Bridge-CSV] file={event_csv.name} processed={i} kept={len(rows)} dropped={dropped}", flush=True)

        note = f"bridge CSV mode: {unlabeled} event(s) unlabeled in {bridge_label_map_path}" if unlabeled > 0 else None

    ensure_parent(out_bridge)
    with out_bridge.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADER)
        writer.writeheader()
        writer.writerows(rows)

    append_rows(out_all, rows)
    return len(rows), dropped, class_counts, note


def write_placeholder_csv(path: Path) -> None:
    ensure_parent(path)
    if path.exists():
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["note"])
        writer.writerow(["placeholder: source-specific preparation not yet implemented in core script"])


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare and harmonize source data")
    parser.add_argument("--data-root", default="data")
    parser.add_argument(
        "--sources",
        nargs="+",
        default=["tmd", "uci", "bridge"],
        choices=["tmd", "uci", "bridge"],
        help="Choose which sources to process",
    )
    parser.add_argument(
        "--bridge-label-map",
        default="configs/bridge_event_labels.csv",
        help="CSV mapping bridge event timestamps to targets (human/vehicle/noise)",
    )
    parser.add_argument("--progress-every", type=int, default=5000, help="Print progress every N records/rows")
    args = parser.parse_args()

    data_root = Path(args.data_root)
    selected = set(args.sources)
    # Start canonical file fresh with only header.
    write_header(data_root / "interim" / "all_sources_canonical.csv")
    stale_sismi = data_root / "interim" / "sismi_mapped.csv"
    if stale_sismi.exists():
        stale_sismi.unlink()

    tmd_kept, tmd_dropped, tmd_counts = 0, 0, {"human": 0, "vehicle": 0, "noise": 0}
    uci_kept, uci_dropped, uci_counts = 0, 0, {"human": 0, "vehicle": 0, "noise": 0}
    bridge_kept, bridge_dropped, bridge_counts, bridge_note = 0, 0, {"human": 0, "vehicle": 0, "noise": 0}, None

    if "tmd" in selected:
        tmd_kept, tmd_dropped, tmd_counts = prepare_tmd(data_root, args.progress_every)
    if "uci" in selected:
        uci_kept, uci_dropped, uci_counts = prepare_uci_har(data_root, args.progress_every)
    if "bridge" in selected:
        bridge_kept, bridge_dropped, bridge_counts, bridge_note = prepare_bridge(
            data_root,
            args.progress_every,
            Path(args.bridge_label_map),
        )

    print("Stage 1 summary")
    print(f"TMD: kept={tmd_kept}, dropped={tmd_dropped}, counts={tmd_counts}")
    print(f"UCI: kept={uci_kept}, dropped={uci_dropped}, counts={uci_counts}")
    print(f"Bridge: kept={bridge_kept}, dropped={bridge_dropped}, counts={bridge_counts}")
    if bridge_note:
        print(f"Bridge note: {bridge_note}")
    print(f"Wrote canonical: {data_root / 'interim' / 'all_sources_canonical.csv'}")


if __name__ == "__main__":
    main()
